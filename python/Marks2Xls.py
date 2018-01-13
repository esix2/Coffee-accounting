import numpy as np
import sys
import pandas as pd
import time
import os
import shutil
import openpyxl as px
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Border, Alignment, Protection, Font
from openpyxl.cell import get_column_letter
from openpyxl.cell import Cell
from fnmatch import fnmatch

def Marks2Xls():
	os.system("clear")
        pwd = os.getcwd()
        if os.path.isdir("../Marks"):  ### checks if any Mark has been counted already or not
                pass
        else:
                os.makedirs("../Marks")

        Files  = filter(lambda x: fnmatch(x,'Marks_*'),os.listdir(pwd))
        for idx in range(len(Files)):
                csvFile = Files[idx]
		df = pd.read_csv(csvFile)
		xlsFile = "../Marks/"+csvFile+'.xlsx'
		writer = pd.ExcelWriter(xlsFile,engine='xlsxwriter')
		df.to_excel(writer,sheet_name='Sheet1')
		wb = writer.book
		ws = writer.sheets['Sheet1']
		writer.save()
		###### Now the xlsx file must be loaded for cell adjusment using openpyxl
		##########################################################################
		wb = load_workbook(xlsFile)
		ws = wb.get_active_sheet()
		Columns = ws.get_highest_column()
		Rows =  ws.get_highest_row()
	       
		for i in range(1,Rows+1):
			ws.row_dimensions[i].height = 30
		ws.column_dimensions[get_column_letter((1))].width = 5
		ws.column_dimensions[get_column_letter((2))].width = 30
		for i in range(3,Columns+1):
			ws.column_dimensions[get_column_letter((i))].width = 14
		Style_CountMarks(ws,1,1,2,Columns,'lgr')
		Style_CountMarks(ws,2,Rows-1,1,2,'g')
		Style_CountMarks(ws,2,Rows,3,Columns-1,'nb')
		Style_CountMarks(ws,Rows,Rows,1,Columns-1,'y')
		ws.cell(row=0,column=Columns-1).value = "Total ("+u"\u20AC"+")" 
		Style_CountMarks(ws,1,Rows-1,Columns,Columns,'lb')
		wb.save(xlsFile)
#	os.system("soffice " + xlsFile + " &")
def Style_CountMarks(ws,Row_s,Row_e,Col_s,Col_e,c):
        for i in range(Row_s,Row_e+1):
                for j in range(Col_s,Col_e+1):
                #       print "i="+str(i)+", j="+str(j)
                        tmp_cell = ws.cell(row=i-1,column=j-1)
                        if c == 'nb':  ## stands for no background
                                pass
                        else:
                                tmp_cell.style.fill.fill_type = Fill.FILL_SOLID
                                if c == 'g': ## stands for green
                                        tmp_cell.style.fill.start_color.index = Color.GREEN
                                elif c == 'r': ## stands for red
                                        tmp_cell.style.fill.start_color.index = Color.RED
                                elif c == 'lb': ## stands for red
                                        tmp_cell.style.fill.start_color.index = 'FF89C4FF'#'FF6CB5FF'#'FF5CADFF'
                                elif c == 'y': ## stands for yellow
                                        tmp_cell.style.fill.start_color.index = Color.YELLOW
                                elif c == 'o': ## stands for orange
                                        tmp_cell.style.fill.start_color.index = 'FFF4C99F'#'FFEDAB69'
                                elif c == 'lgr': ## stands for light gray
                                        tmp_cell.style.fill.start_color.index = 'FFE6E6E6'
                                elif c == 'dgr': ## stands for dark gray
                                        tmp_cell.style.fill.start_color.index = 'FFBABABA'
			if i == 1 or j == 1 or j == 2:
				tmp_cell.style.font.name = 'Times New Roman'
				tmp_cell.style.font.bold = True
			else:
				tmp_cell.style.font.name = 'Arial'
				tmp_cell.style.font.bold = False
                        tmp_cell.style.font.size = 12
                        tmp_cell.style.alignment.vertical = "center"
			tmp_cell.style.alignment.horizontal = "center"
                        tmp_cell.style.alignment.wrap_text = True
                        tmp_cell.style.borders.top.border_style = Border.BORDER_THIN
                        tmp_cell.style.borders.bottom.border_style = Border.BORDER_THIN
                        tmp_cell.style.borders.right.border_style = Border.BORDER_THIN
                        tmp_cell.style.borders.left.border_style = Border.BORDER_THIN
       
