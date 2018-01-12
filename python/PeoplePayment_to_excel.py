import pandas as pd
from openpyxl.styles import Color, Fill, Border, Alignment, Font
from openpyxl.cell import get_column_letter
from openpyxl.comments import Comment
import numpy as np
def PeoplePayment_to_excel():
        csvFile = 'PeoplePayment'
        df = pd.read_csv(csvFile)
        df.index = range(1,len(df)+1)
        xlsFile = csvFile+'.xlsx'
        writer = pd.ExcelWriter(xlsFile,engine='openpyxl')
        Header_main = df.columns.values
        Header_payment = Header_main[range(3,len(Header_main))];
        main_idx = range(0,3)
	main_idx.extend(range(3,len(Header_main),2))
	Dates  = df[Header_payment[range(1,len(Header_payment),2)]]
	df  = df[Header_main[main_idx]]

        csvFile = 'PeopleBalance'
        tmp = pd.read_csv(csvFile)
        tmp.index = range(1,len(df)+1)
        balance = tmp['Balance']
        df['Balance'] = balance

        df.to_excel(writer,sheet_name='Sheet1')
        wb = writer.book
        ws = writer.sheets['Sheet1']
        Columns = ws.get_highest_column()
        Rows =  ws.get_highest_row()
	ws.row_dimensions[1].height = 50
        for i in range(2,Rows+1):
                ws.row_dimensions[i].height = 30
        ws.column_dimensions[get_column_letter((1))].width = 5
        ws.column_dimensions[get_column_letter((2))].width = 25
        for j in range(3,Columns+1):
		ws.column_dimensions[get_column_letter((j))].width = 15

	Style_MembersPayment(ws,1,1,2,Columns,'dgr')
	Style_MembersPayment(ws,2,Rows,1,2,'g')
	Style_MembersPayment(ws,2,Rows,3,Columns,'nb')
        Style_MembersPayment(ws,2,Rows,Columns,Columns,'dgr')
	idx = 0 
	##############################################################
	############ Insert payment date as the comment into the cells
	############################################################## 
	for j in range(4,Columns-1):
		tmp = Dates[Dates.columns.values[idx]].values
		tmp_cell = ws.cell(row=0,column=j)
		tmp_cell.value = ''
		tmp_cell.style.borders.diagonal.border_style = Border.BORDER_DOTTED
		for i in range(1,Rows):
			tmp_cell = ws.cell(row=i,column=j)
			if tmp[i-1] > 0:
				comm = Comment(tmp[i-1], 'Ehsan Zandi')
				tmp_cell.comment = comm
			else:
				pass
		idx += 1
        ws['C1'].value = "Last Termination\n ("+u"\u20AC"+")"
        ws['D1'].value = "Sum\n ("+u"\u20AC"+")"
        ws.cell(row=0,column=Columns-1).value = "Balance\n ("+u"\u20AC"+")"
	return ws
def Style_MembersPayment(ws,Row_s,Row_e,Col_s,Col_e,c):
        for i in range(Row_s,Row_e+1):
                for j in range(Col_s,Col_e+1):
                #       print "i="+str(i)+", j="+str(j)
                        tmp_cell = ws.cell(row=i-1,column=j-1)
                        if ws.cell(row=0,column=j-1).value == 'Balance':  
				if tmp_cell.value < 0:
					tmp_cell.style.fill.fill_type = Fill.FILL_SOLID
					tmp_cell.style.fill.start_color.index = Color.RED
				else:
					tmp_cell.style.fill.start_color.index = Color.GREEN
                        else:
				if c == 'nb':  ## stands for no background
					pass
				else:
					tmp_cell.style.fill.fill_type = Fill.FILL_SOLID
					if c == 'g': ## stands for green
						tmp_cell.style.fill.start_color.index = Color.GREEN
					elif c == 'r': ## stands for red
						tmp_cell.style.fill.start_color.index = Color.RED
					elif c == 'lb': ## stands for red
						tmp_cell.style.fill.start_color.index = 'FF89C4FF'#'FF6CB5FF'#'FF5CADFF'
					elif c == 'y': ## stands for yellow
						tmp_cell.style.fill.start_color.index = Color.YELLOW
					elif c == 'o': ## stands for orange
						tmp_cell.style.fill.start_color.index = 'FFF4C99F'#'FFEDAB69'
					elif c == 'lgr': ## stands for light gray
						tmp_cell.style.fill.start_color.index = 'FFE6E6E6'
					elif c == 'dgr': ## stands for dark gray
						tmp_cell.style.fill.start_color.index = 'FFBABABA'
                        if i == 1 or j == 1 or j == 2:
                                tmp_cell.style.font.name = 'Times New Roman'
                                tmp_cell.style.font.bold = True
                        else:
                                tmp_cell.style.font.name = 'Arial'
                                tmp_cell.style.font.bold = False
                        tmp_cell.style.font.size = 12
                        tmp_cell.style.alignment.vertical = "center"
                        tmp_cell.style.alignment.horizontal = "center"
                        tmp_cell.style.alignment.wrap_text = True
                        tmp_cell.style.borders.top.border_style = Border.BORDER_THIN
                        tmp_cell.style.borders.bottom.border_style = Border.BORDER_THIN
                        tmp_cell.style.borders.right.border_style = Border.BORDER_THIN
                        tmp_cell.style.borders.left.border_style = Border.BORDER_THIN
