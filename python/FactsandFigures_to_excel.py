import pandas as pd
import os
from openpyxl.styles import Color, Fill, Border, Alignment, Font
from openpyxl.cell import get_column_letter
from FactsandFigures import FactsandFigures
def FactsandFigures_to_excel():
        csvFile = 'FactsandFigures'
 df = pd.read_csv(csvFile)
        df = df.T
        Index = ['Last Version Benefit','Total Income','Current Cash', 'System Expenditure','Inventory Value','Members Credit','System Balance','Security']
        xlsFile = csvFile+'.xlsx'
        writer = pd.ExcelWriter(xlsFile,engine='openpyxl')
        df.to_excel(writer,sheet_name='Sheet1')

        wb = writer.book
        ws = writer.sheets['Sheet1']
        Columns = ws.get_highest_column()
        Rows =  ws.get_highest_row()
        for i in range(1,Rows):
                if i == Rows-1:
   ws['A'+str(i)].value = ws['A'+str(i+1)].value+" (%)"
                else:
   ws['A'+str(i)].value = ws['A'+str(i+1)].value+"("+u"\u20AC"+")"
                ws['B'+str(i)].value = ws['B'+str(i+1)].value
 ws['A'+str(Rows)] = ''
 ws['B'+str(Rows)] = ''
        for i in range(1,Rows):
                ws.row_dimensions[i].height = 50
        ws.column_dimensions[get_column_letter((1))].width = 35
        ws.column_dimensions[get_column_letter((2))].width = 15

        Style_Facts(ws,1,Rows-1,1,1,'y')
        Style_Facts(ws,1,Rows-1,2,Columns,'nb')

 return ws
def Style_Facts(ws,Row_s,Row_e,Col_s,Col_e,c):
        for i in range(Row_s,Row_e+1):
                for j in range(Col_s,Col_e+1):
                #       print "i="+str(i)+", j="+str(j)
                        tmp_cell = ws.cell(row=i-1,column=j-1)
                        if c == 'nb':  ## stands for no background
                                pass
                        else:
                                tmp_cell.style.fill.fill_type = Fill.FILL_SOLID
                                if c == 'g': ## stands for green
                                        tmp_cell.style.fill.start_color.index = Color.GREEN
                                elif c == 'r': ## stands for red
                                        tmp_cell.style.fill.start_color.index = Color.RED
                                elif c == 'lb': ## stands for red
                                        tmp_cell.style.fill.start_color.index = 'FF89C4FF'#'FF6CB5FF'#'FF5CADFF'
                                elif c == 'y': ## stands for yellow
                                        tmp_cell.style.fill.start_color.index = Color.YELLOW
                                elif c == 'o': ## stands for orange
                                        tmp_cell.style.fill.start_color.index = 'FFF4C99F'#'FFEDAB69'
                                elif c == 'lgr': ## stands for light gray
                                        tmp_cell.style.fill.start_color.index = 'FFE6E6E6'
                                elif c == 'dgr': ## stands for dark gray
                                        tmp_cell.style.fill.start_color.index = 'FFBABABA'
                        if j == 1:
                                tmp_cell.style.font.name = 'Times New Roman'
                                tmp_cell.style.font.bold = True
                        else:
                                tmp_cell.style.font.name = 'Arial'
                                tmp_cell.style.font.bold = False
                        tmp_cell.style.alignment.vertical = "center"
   tmp_cell.style.alignment.horizontal = "center"
                        tmp_cell.style.alignment.wrap_text = True
                        tmp_cell.style.borders.top.border_style = Border.BORDER_THIN
                        tmp_cell.style.borders.bottom.border_style = Border.BORDER_THIN
                        tmp_cell.style.borders.right.border_style = Border.BORDER_THIN
                        tmp_cell.style.borders.left.border_style = Border.BORDER_THIN
       
