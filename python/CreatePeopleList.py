import numpy as np
import sys
import pandas as pd
import time
import os
import shutil
import openpyxl as px
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Border, Alignment, Protection, Font
from openpyxl.cell import get_column_letter
from openpyxl.cell import Cell
def CreatePeopleList(addListofDrinksatBottom):
	xlsFile = os.path.dirname(os.getcwd())+'/AllPeople.xlsx'
	pdfFile =  os.path.dirname(os.getcwd())+'AllPeople.pdf'
        pwd = os.getcwd()
        df = pd.read_csv('PeopleList')
        PeopleNames = df['Name']
	tmp = np.empty((2*len(PeopleNames)+5,1))
	tmp[:] = np.nan
	tmp = pd.DataFrame(tmp)
	idx_displacement = 0
	tmp.loc[0,0] = 'Name'
	for i in range(0,len(PeopleNames)):
		if PeopleNames[i] == 'FFTI':
			idx_displacement = -2 
			i = i+1
		else:
			tmp.loc[2*i+idx_displacement+1,0] = PeopleNames[i]	
	tmp.loc[2*len(PeopleNames)-1,0] = 'Student Assistants'
	tmp.loc[2*len(PeopleNames)+3,0] = 'Guests'
	People = tmp[0]
        os.chdir("../../../../python")
        df = pd.read_csv('CoffeeList',sep='\t')
        os.chdir(pwd)
        DrinkNames = df['Drink']

	tmp = np.empty((3*len(DrinkNames),1))
	tmp[:] = np.nan
	tmp = pd.DataFrame(tmp)
	for i in range(0,len(DrinkNames)):
		tmp.loc[3*i,0] = DrinkNames[i]	
		tmp.loc[3*i+1,0] = DrinkNames[i]	
		tmp.loc[3*i+2,0] = DrinkNames[i]	
	Drinks = tmp[0]
        NumberofDrinks =  len(Drinks)
        NumberofPeople = len(People)
        Table = np.empty((NumberofPeople,NumberofDrinks))
        Table[:] = np.nan
        Table = pd.DataFrame(Table,index=People,columns=Drinks)
#       print Table

        writer = pd.ExcelWriter(xlsFile,engine='xlsxwriter')
        Table.to_excel(writer,sheet_name='Sheet1')
        wb = writer.book
        ws = writer.sheets['Sheet1']
        writer.save()
	##########################################################################
	###### Now the xlsx file must be loaded for cell adjusment using openpyxl
	##########################################################################
	wb = load_workbook(xlsFile)
        ws = wb.get_active_sheet()
	Columns = ws.get_highest_column()
	Rows =  ws.get_highest_row() 
	ws.page_setup.paperSize = ws.PAPERSIZE_A3
	ws.page_setup.fitToPage = True
	ws.page_setup.fitToHeight = 0
	ws.page_setup.fitToWidth = 1
	ws.page_setup.horizontalCentered = True
	ws.page_setup.verticalCentered = True	
	ws.cell(coordinate='C1').value = ''
	ws.cell(coordinate='D1').value = ''
	ws.merge_cells('B1:D2')
	ws.merge_cells('E1:G2')
	ws.merge_cells('H1:J2')
	ws.merge_cells('K1:M2')
	ws.merge_cells('N1:P2')
	ws.merge_cells('Q1:S2')
	ws.merge_cells('T1:V2')
	ws.merge_cells('W1:Y2')
	i = 3
        while 3 <= i < Rows:
		if ws.cell(coordinate="A"+str(i)).value == 'Student Assistants':
			ws.merge_cells("A"+str(i)+":A"+str(i+3))
			i += 4
		elif ws.cell(coordinate="A"+str(i)).value == 'Guests':
			ws.merge_cells("A"+str(i)+":A"+str(i+1))
			i += 2
		else:
			ws.merge_cells("A"+str(i)+":A"+str(i+1))
			i += 2
	h1 = 20 ## hieght of the first row
	w1 = 30
	tmargin = 19;
	rmargin = 18;
	A3_h = 1050
	A3_w = 150
	tmargin = 0.1*A3_h
	rmargin = 0.1*A3_w
	h = (A3_h-(0*tmargin)-2*h1)/(Rows-2)
	w = (A3_w-(0*rmargin)-w1)/(Columns-1)
#	print get_column_letter((7))
#	ws.column_dimensions[get_column_letter((7))].width = 50 
	ws.cell(row=0, column=0).value = 'Date:'
        for i in range(1,2):
		ws.row_dimensions[i].height = h1
        for i in range(3,Rows+1):
		ws.row_dimensions[i].height = h
	for j in range(1,2):
		ws.column_dimensions[get_column_letter((j))].width = w1
	for j in range(2,Columns+1):
		ws.column_dimensions[get_column_letter((j))].width = w
	Sytle_CreatePeopleList(ws,1,1,1,1,w1,h1,'Times New Roman','nb',Rows)
	Sytle_CreatePeopleList(ws,2,2,1,1,w1,h1,'Times New Roman','lgr',Rows)
	Sytle_CreatePeopleList(ws,1,2,2,Columns,w1,h1,'Times New Roman','nb',Rows)
	myColor = 'lgr'
	i = 3
        while 3 <= i < Rows: ## sets the color of every pair of  consequent rows to light and dark gray 
		if myColor == 'lgr':
			myColor = 'dgr'
		else:
			myColor = 'lgr'
		if ws.cell(coordinate="A"+str(i)).value == 'Student Assistants':
			Sytle_CreatePeopleList(ws,i,i+3,1,Columns,w,h,'Times New Roman','o',Rows)
			i += 4
		elif ws.cell(coordinate="A"+str(i)).value == 'Guests':
			Sytle_CreatePeopleList(ws,i,i+1,1,Columns,w,h,'Times New Roman','lb',Rows)
			i += 2
		else:
			Sytle_CreatePeopleList(ws,i,i+1,1,Columns,w,h,'Times New Roman',myColor,Rows)
			i += 2
	ws.cell(row=0,column=0).style.borders.bottom.border_style = Border.BORDER_THIN
	ws.cell(row=1,column=0).style.borders.top.border_style = Border.BORDER_THIN 
	if addListofDrinksatBottom == True:  ##### Adds the list of drink one more time at the bottom of the list
		for k in range(0,2):
			for j in range(0,len(DrinkNames)):
				last_cell = ws.cell(row=Rows+k, column=3*j+1)
				ws.cell(row=Rows+k, column=3*j+1).value = DrinkNames[j]
				ws.cell(row=Rows+k, column=3*j+2).value = DrinkNames[j]
				ws.cell(row=Rows+k, column=3*j+3).value = DrinkNames[j]
		ws.merge_cells("B"+str(Rows+1)+":D"+str(Rows+2))
		ws.merge_cells("E"+str(Rows+1)+":G"+str(Rows+2))
		ws.merge_cells("H"+str(Rows+1)+":J"+str(Rows+2))
		ws.merge_cells("K"+str(Rows+1)+":M"+str(Rows+2))
		ws.merge_cells("N"+str(Rows+1)+":P"+str(Rows+2))
		ws.merge_cells("Q"+str(Rows+1)+":S"+str(Rows+2))
		ws.merge_cells("T"+str(Rows+1)+":V"+str(Rows+2))
		ws.merge_cells("W"+str(Rows+1)+":W"+str(Rows+2))
		Sytle_CreatePeopleList(ws,Rows+1,Rows+2,2,Columns,w1,h1,'Times New Roman','nb',Rows)
		for j in range(0,len(DrinkNames)):
			ws.cell(row=Rows+1,column=3*j+1).style.borders.bottom.border_style = Border.BORDER_THIN
			ws.cell(row=Rows,column=3*j+1).style.borders.top.border_style = Border.BORDER_THIN
        wb.save(xlsFile)
        os.chdir("../")
	os.system("libreoffice --headless --convert-to pdf "+ xlsFile)
	os.remove(xlsFile)
        os.chdir(pwd)
#	os.system("soffice " + xlsFile + " &")

def Sytle_CreatePeopleList(ws,Row_s,Row_e,Col_s,Col_e,w,h,myFont,c,Rows):
	leftBorder  = ['A','B','E','H','K','N','Q','T','W']
	rightBorder = ['A','D','G','J','M','P','S','V','Y']
	topBorder = range(1,Rows,2)
#	print topBorder
	#print Rows
	#print topBorder.index(Rows-3)
	#print Rows-3
	#topBorder.remove(topBorder.index(Rows-3))
	bottomBorder = range(2,Rows+1,2)
	#bottomBorder.remove(bottomBorder.index(Rows-4))
        for i in range(Row_s,Row_e+1):
		if (i in topBorder) and (i != Rows-3):
			tstyle = Border.BORDER_THIN 
		else:
			tstyle = Border.BORDER_DOTTED
		if (i in bottomBorder) and (i != Rows -4):
				bstyle = Border.BORDER_THIN 
		else:
			bstyle = Border.BORDER_DOTTED
                for j in range(Col_s,Col_e+1):
		#	print "i="+str(i)+", j="+str(j)
                        tmp_cell = ws.cell(row=i-1,column=j-1)
                        if c == 'nb':  ## stands for no background
				pass 
			else:
				tmp_cell.style.fill.fill_type = Fill.FILL_SOLID
				if c == 'g': ## stands for green
					tmp_cell.style.fill.start_color.index = Color.GREEN
				elif c == 'r': ## stands for red
					tmp_cell.style.fill.start_color.index = Color.RED
				elif c == 'lb': ## stands for red
					tmp_cell.style.fill.start_color.index = 'FF89C4FF'#'FF6CB5FF'#'FF5CADFF'
				elif c == 'y': ## stands for yellow
					tmp_cell.style.fill.start_color.index = Color.YELLOW
				elif c == 'o': ## stands for orange
					tmp_cell.style.fill.start_color.index = 'FFF4C99F'#'FFEDAB69'
				elif c == 'lgr': ## stands for light gray
					tmp_cell.style.fill.start_color.index = 'FFE6E6E6'
				elif c == 'dgr': ## stands for dark gray
					tmp_cell.style.fill.start_color.index = 'FFBABABA'
                        tmp_cell.style.font.name = myFont
                        tmp_cell.style.font.size = 12
                        tmp_cell.style.font.bold = True
                        tmp_cell.style.alignment.vertical = "center"
			if i==1 and j==1:
				tmp_cell.style.alignment.horizontal = "left"
			else:
				tmp_cell.style.alignment.horizontal = "center"
                        tmp_cell.style.alignment.wrap_text = True
			if get_column_letter((j)) in leftBorder:
				lstyle = Border.BORDER_THIN 
			else:
				lstyle = Border.BORDER_DOTTED
			if get_column_letter((j)) in rightBorder:
				rstyle = Border.BORDER_THIN 
			else:
				rstyle = Border.BORDER_DOTTED
                        tmp_cell.style.borders.top.border_style = tstyle
                        tmp_cell.style.borders.bottom.border_style = bstyle
                        tmp_cell.style.borders.left.border_style = lstyle
                        tmp_cell.style.borders.right.border_style = rstyle
         #               if i == Row_s:
         #                       ws.column_dimensions[get_column_letter((j))].width = w

