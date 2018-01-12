import pandas as pd
import os
from openpyxl.styles import Color, Fill, Border, Alignment, Font
from openpyxl.cell import get_column_letter
import numpy as np
def ShoppingList_to_excel():
        pwd = os.getcwd()
        os.chdir("../..") ## it changes to the parent folder, since shopping list is there
        csvFile = 'ShoppingList'
        df = pd.read_csv(csvFile)
        if float(sum(1 for line in open(csvFile))) < 2:
   df = pd.DataFrame(np.nan, index=[0], columns=df.columns)
        xlsFile = csvFile+'.xlsx'
        writer = pd.ExcelWriter(xlsFile,engine='openpyxl')
        df.index = range(1,len(df)+1)
        df.to_excel(writer,sheet_name='Sheet1')

        wb = writer.book
        ws = writer.sheets['Sheet1']
        Columns = ws.get_highest_column()
        Rows =  ws.get_highest_row()

 ws.row_dimensions[1].height = 50
        for i in range(2,Rows+1):
                ws.row_dimensions[i].height = 30
        ws.column_dimensions[get_column_letter((1))].width = 5
        ws.column_dimensions[get_column_letter((2))].width = 15
        ws.column_dimensions[get_column_letter((3))].width = 30
        ws.column_dimensions[get_column_letter((4))].width = 10
        ws.column_dimensions[get_column_letter((5))].width = 10
        ws.column_dimensions[get_column_letter((6))].width = 15
        ws.column_dimensions[get_column_letter((7))].width = 10
        ws.column_dimensions[get_column_letter((8))].width = 15
        ws.column_dimensions[get_column_letter((9))].width = 15
        ws.column_dimensions[get_column_letter((10))].width = 10
        ws.column_dimensions[get_column_letter((11))].width = 15
        Style_ShoppingList(ws,1,Rows,1,Columns,'nb')

        Style_ShoppingList(ws,1,1,2,Columns,'g')
        Style_ShoppingList(ws,2,Rows,1,1,'g')
        ws['F1'].value = "Price Per item ("+u"\u20AC"+")"
        ws['G1'].value = "Tip\n ("+u"\u20AC"+")"
        ws['H1'].value = "Total Price\n ("+u"\u20AC"+")"

        os.chdir(pwd)
 return ws
def Style_ShoppingList(ws,Row_s,Row_e,Col_s,Col_e,c):
        for i in range(Row_s,Row_e+1):
                for j in range(Col_s,Col_e+1):
                #       print "i="+str(i)+", j="+str(j)
                        tmp_cell = ws.cell(row=i-1,column=j-1)
                        if c == 'nb':  ## stands for no background
                                pass
                        else:
                                tmp_cell.style.fill.fill_type = Fill.FILL_SOLID
                                if c == 'g': ## stands for green
                                        tmp_cell.style.fill.start_color.index = Color.GREEN
                                elif c == 'r': ## stands for red
                                        tmp_cell.style.fill.start_color.index = Color.RED
                                elif c == 'lb': ## stands for red
                                        tmp_cell.style.fill.start_color.index = 'FF89C4FF'#'FF6CB5FF'#'FF5CADFF'
                                elif c == 'y': ## stands for yellow
                                        tmp_cell.style.fill.start_color.index = Color.YELLOW
                                elif c == 'o': ## stands for orange
                                        tmp_cell.style.fill.start_color.index = 'FFF4C99F'#'FFEDAB69'
                                elif c == 'lgr': ## stands for light gray
                                        tmp_cell.style.fill.start_color.index = 'FFE6E6E6'
                                elif c == 'dgr': ## stands for dark gray
                                        tmp_cell.style.fill.start_color.index = 'FFBABABA'
                        if i == 1 or j == 1:
                                tmp_cell.style.font.name = 'Times New Roman'
                                tmp_cell.style.font.bold = True
                        else:
                                tmp_cell.style.font.name = 'Arial'
                                tmp_cell.style.font.bold = False
                        tmp_cell.style.alignment.vertical = "center"
   tmp_cell.style.alignment.horizontal = "center"
                        tmp_cell.style.alignment.wrap_text = True
                        tmp_cell.style.borders.top.border_style = Border.BORDER_THIN
                        tmp_cell.style.borders.bottom.border_style = Border.BORDER_THIN
                        tmp_cell.style.borders.right.border_style = Border.BORDER_THIN
                        tmp_cell.style.borders.left.border_style = Border.BORDER_THIN
       
